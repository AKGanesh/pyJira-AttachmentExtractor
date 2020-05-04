# -*- coding: utf-8 -*-
"""
Created on Mon Mar 30 17:08:13 2020
JIRA - CV extractor based on JQL

@author: Ganesh Akondi
Spyder - Py 3.7
"""

import os.path
import shutil
from jira.client import JIRA
from openpyxl import Workbook

#JIRA server URL and auth
jira_options={'server': 'https://<company>.atlassian.net'}
jira=JIRA(options=jira_options,basic_auth=('email@domain.com','accesstoken'))

#Directory to save extracted attachments
dir_to_save = 'D:\\jira-plugin\\CV'

#Excel workbook for generating summary of attachments with Hyperlinks to open the attachments
wb = Workbook()
ws = wb.active
key_row = 1
parent_key_row = 1
summary_row = 1
description_row = 1
attachment_row = 1

start_column = 1

#JQL Query - Ref: https://www.atlassian.com/blog/jira-software/jql-the-most-flexible-way-to-search-jira-14 
jql_query = 'project = REC AND issuetype in (Epic, Story, Sub-task) AND "Epic Link" not in (REC-XX, REC-XX) AND Sprint in openSprints()'
stories_in_project = jira.search_issues(jql_query, fields='*all')

for story in stories_in_project:
    folder_name = "\\"+story.key+"_"+story.fields.summary.strip()
    save_path = dir_to_save + folder_name
    if os.path.exists(save_path):
        shutil.rmtree(save_path)
    os.makedirs(save_path)

    for subtask in story.fields.subtasks:
        issue = jira.issue(subtask.key, fields='*all')
        for attachment in issue.fields.attachment:
            if type(attachment) == str:
                break
            else:
                name_of_file = attachment.filename
                completeName = os.path.join(save_path, name_of_file)
                print(completeName)
                footage = attachment.get()
                with open(completeName, 'wb') as f:        
                    f.write(footage)
                    
        ws.cell(row=parent_key_row, column=start_column).value = folder_name
        parent_key_row += 1
        
        ws.cell(row=key_row, column=start_column+1).value = issue.key
        key_row += 1
        
        ws.cell(row=description_row, column=start_column+3).value = issue.fields.description
        description_row += 1
        
        ws.cell(row=attachment_row, column=start_column+2).value = attachment.filename
        ws.cell(row=attachment_row, column=start_column+2).hyperlink = completeName
        ws.cell(row=attachment_row, column=start_column+2).style = "Hyperlink"
        attachment_row += 1
        
# set the width of the column 
ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 50
ws.column_dimensions['D'].width = 140

#Save the excel file    
wb.save("D:\\jira-plugin\\jira-report.xlsx")

print('-- Execution completed, please check the folder --')